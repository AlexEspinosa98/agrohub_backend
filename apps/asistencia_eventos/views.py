import io
import json
import uuid

import pandas as pd
from django.core.files.storage import default_storage
from django.http import HttpResponse
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from rest_framework import status
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.exceptions import NotFound, ParseError
from rest_framework.response import Response

from apps.asistencia_eventos import services
from apps.asistencia_eventos.models import Evento, PersonaAsistente
from apps.asistencia_eventos.ocr_service import extract_asistencia
from apps.asistencia_eventos.serializers import EventoConfirmSerializer
from apps.user_activity.authentication import TokenHeaderAuthentication
from apps.user_activity.permissions import IsAdminRole, IsAuthenticatedWithRole

_AUTH = [TokenHeaderAuthentication]
_ADMIN_ONLY = [IsAuthenticatedWithRole, IsAdminRole]


def _save_scan(upload_file) -> str:
    filename = f"{uuid.uuid4()}_{upload_file.name}"
    return default_storage.save(f"asistencia_eventos/{filename}", upload_file)


@api_view(["POST"])
@authentication_classes(_AUTH)
@permission_classes(_ADMIN_ONLY)
def scan_evento(request):
    """Paso 1: sube el PDF/imagen escaneado, corre OCR y devuelve TODO lo
    extraído (encabezado del evento + asistentes + texto crudo del OCR) para
    que se revise/corrija en el flujo web antes de guardar nada."""
    upload = request.FILES.get("archivo")
    if not upload:
        raise ParseError("Falta el archivo escaneado (campo 'archivo')")

    extracted = extract_asistencia(upload)

    documentos = [a["numero_documento"] for a in extracted["asistentes"] if a.get("numero_documento")]
    existentes = set(
        PersonaAsistente.objects.filter(numero_documento__in=documentos).values_list(
            "numero_documento", flat=True
        )
    )
    for asistente in extracted["asistentes"]:
        asistente["persona_ya_registrada"] = asistente.get("numero_documento") in existentes

    return Response(
        {
            "status": status.HTTP_200_OK,
            "message": "Documento escaneado — revisa y corrige antes de guardar",
            "data": extracted,
        }
    )


def _crear_evento(request):
    """Paso 2: recibe la información ya depurada (misma forma que el draft
    del scan) junto con el archivo original, y ahí sí guarda todo — sube el
    documento escaneado, hace upsert de cada persona por número de
    documento (garantizando unicidad) y crea el evento + sus registros de
    asistencia.

    Multipart no anida JSON dentro de campos de formulario, así que cuando
    va acompañado del archivo, el payload depurado viaja como un campo
    'data' con el JSON como texto; si se manda sin archivo (application/json
    puro) se acepta el body tal cual."""
    if "data" in request.data:
        try:
            payload = json.loads(request.data["data"])
        except (TypeError, ValueError) as exc:
            raise ParseError("El campo 'data' debe ser un JSON válido") from exc
    else:
        payload = request.data

    serializer = EventoConfirmSerializer(data=payload)
    serializer.is_valid(raise_exception=True)
    data = serializer.validated_data

    upload = request.FILES.get("archivo")
    documento_path = _save_scan(upload) if upload else None
    texto_crudo_ocr = payload.get("texto_crudo_ocr") or None

    evento = services.guardar_evento(data, documento_path, texto_crudo_ocr)
    return Response(
        {
            "status": status.HTTP_201_CREATED,
            "message": "evento guardado",
            "data": {"id": evento.id, "total_asistentes": evento.asistentes.count()},
        },
        status=status.HTTP_201_CREATED,
    )


def _listar_eventos(request):
    eventos = Evento.objects.order_by("-fecha", "-id")
    data = [
        {
            "id": e.id,
            "tema": e.tema,
            "responsable": e.responsable,
            "lugar": e.lugar,
            "fecha": e.fecha,
            "total_asistentes": e.asistentes.count(),
        }
        for e in eventos
    ]
    return Response({"status": status.HTTP_200_OK, "message": "eventos", "data": data})


@api_view(["GET", "POST"])
@authentication_classes(_AUTH)
@permission_classes(_ADMIN_ONLY)
def eventos_list_create(request):
    if request.method == "POST":
        return _crear_evento(request)
    return _listar_eventos(request)


@api_view(["GET"])
@authentication_classes(_AUTH)
@permission_classes(_ADMIN_ONLY)
def evento_detail(request, evento_id: int):
    evento = Evento.objects.filter(id=evento_id).first()
    if not evento:
        raise NotFound("Evento no encontrado")

    asistentes = [
        {
            "nombre": r.persona.nombre,
            "tipo_documento": r.persona.tipo_documento,
            "numero_documento": r.persona.numero_documento,
            "genero": r.persona.genero,
            "pertenencia_etnica": r.persona.pertenencia_etnica,
            "municipio": r.municipio,
            "telefono": r.telefono,
            "edad": r.edad,
        }
        for r in evento.asistentes.select_related("persona").all()
    ]
    return Response(
        {
            "status": status.HTTP_200_OK,
            "message": "evento",
            "data": {
                "id": evento.id,
                "tema": evento.tema,
                "responsable": evento.responsable,
                "lugar": evento.lugar,
                "fecha": evento.fecha,
                "hora_inicio": evento.hora_inicio,
                "hora_final": evento.hora_final,
                "documento_escaneado": (
                    default_storage.url(evento.documento_escaneado) if evento.documento_escaneado else None
                ),
                "asistentes": asistentes,
            },
        }
    )


@api_view(["GET"])
@authentication_classes(_AUTH)
@permission_classes(_ADMIN_ONLY)
def dashboard_resumen(request):
    return Response({"status": status.HTTP_200_OK, "message": "resumen", "data": services.resumen_general()})


@api_view(["GET"])
@authentication_classes(_AUTH)
@permission_classes(_ADMIN_ONLY)
def dashboard_estadisticas(request):
    evento_id = request.query_params.get("evento_id")
    data = services.estadisticas_por_municipio(evento_id=int(evento_id) if evento_id else None)
    return Response({"status": status.HTTP_200_OK, "message": "estadisticas", "data": data})


_HEADER_FILL = PatternFill(start_color="1F6F43", end_color="1F6F43", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_BOLD_FONT = Font(bold=True)


def _autosize_columns(worksheet):
    for col_cells in worksheet.columns:
        length = max((len(str(c.value)) if c.value is not None else 0) for c in col_cells)
        worksheet.column_dimensions[get_column_letter(col_cells[0].column)].width = min(max(length + 2, 10), 45)


def _formatear_libro(workbook):
    for worksheet in workbook.worksheets:
        if worksheet.max_row == 0:
            continue
        for cell in worksheet[1]:
            cell.font = _HEADER_FONT
            cell.fill = _HEADER_FILL
        worksheet.freeze_panes = "A2"
        _autosize_columns(worksheet)
        if worksheet.title == "Estadisticas" and worksheet.max_row > 1:
            for cell in worksheet[worksheet.max_row]:
                cell.font = _BOLD_FONT


@api_view(["GET"])
@authentication_classes(_AUTH)
@permission_classes(_ADMIN_ONLY)
def dashboard_excel(request):
    """Excel completo del dashboard (opcionalmente filtrado por un solo
    evento con ?evento_id=) — 4 hojas: Resumen, Estadisticas (la misma
    tabla Municipio x Género x edad del dashboard), Eventos y el detalle
    completo de Asistentes."""
    evento_id_param = request.query_params.get("evento_id")
    evento_id = int(evento_id_param) if evento_id_param else None

    resumen = services.resumen_general()
    estadisticas = services.estadisticas_por_municipio(evento_id=evento_id)
    eventos = services.eventos_para_export(evento_id=evento_id)
    asistentes = services.asistentes_para_export(evento_id=evento_id)

    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        pd.DataFrame(
            [
                {"Indicador": "Total personas", "Valor": resumen["total_personas"]},
                {"Indicador": "Total eventos", "Valor": resumen["total_eventos"]},
                {"Indicador": "Total asistencias", "Valor": resumen["total_asistencias"]},
            ]
        ).to_excel(writer, index=False, sheet_name="Resumen")

        columnas_stats = ["municipio", "genero", "0-14", "15-19", "20-59", "mayor de 60", "total"]
        etiquetas_stats = ["Municipio", "Género", "0-14", "15-19", "20-59", "Mayor de 60", "Total"]
        df_stats = pd.DataFrame(estadisticas, columns=columnas_stats)
        df_stats.columns = etiquetas_stats
        df_stats.to_excel(writer, index=False, sheet_name="Estadisticas")

        columnas_eventos = ["id", "tema", "responsable", "lugar", "fecha", "hora_inicio", "hora_final", "total_asistentes"]
        pd.DataFrame(eventos, columns=columnas_eventos).to_excel(writer, index=False, sheet_name="Eventos")

        columnas_asistentes = [
            "evento_id",
            "evento_tema",
            "evento_fecha",
            "nombre",
            "tipo_documento",
            "numero_documento",
            "genero",
            "pertenencia_etnica",
            "municipio",
            "telefono",
            "edad",
        ]
        pd.DataFrame(asistentes, columns=columnas_asistentes).to_excel(
            writer, index=False, sheet_name="Asistentes"
        )

        _formatear_libro(writer.book)

    output.seek(0)
    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = "attachment; filename=dashboard_asistencia_eventos.xlsx"
    return response
